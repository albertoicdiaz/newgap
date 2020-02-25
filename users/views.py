from .utils import *
from .models import *
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import login as do_login
from django.contrib.auth import logout as do_logout
from django.contrib.auth.models import User
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.db.models import Sum, F, Case, When, IntegerField
from django.db.models.functions import Coalesce
import datetime
from datetime import datetime, timedelta
import time
import xlwt
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
import dateutil.parser
import numpy as np
from django.views.decorators.csrf import csrf_exempt
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib
import re

def welcome(request):

    # Si estamos identificados devolvemos la portada
    dominio_num=0
    questions=0
    dom_description = []
    questions_filter = []
    dominio_quantity = []
    dominio_filter = []
    new_doms= []
    dominio_answered = []
    dominios = Dominio.objects.all()
    fechafin=0
    fechainicio=0
    encuesta_sinabrir = 0
    fechasinabrir=0

    
    if request.user.is_authenticated:
        if request.GET.get('dominio'):
            dominio_num = request.GET.get('dominio')
            questions = Pregunta.objects.filter(seccion__dominio__id=dominio_num)

            checkasign = Encuesta.objects.filter(realizar_enc__empleado__usuario__credenciales=request.user.id)
            fechahoy=str(datetime.today().strftime('%Y-%m-%d'))
            # fechahoy="2020-01-22"
            # fechahoy="2020-02-01"
            # fechahoy="2020-02-20"
            count = 0
            for check in checkasign:
                if (fechahoy>=str(check.fecha_inicio) and fechahoy<=str(check.fecha_termino)):
                    count+=1
                    encuesta=check.id
                    for question in questions:
                        count=0
                        count = Respuesta.objects.filter(pregunta__pregunta=question).filter(encuesta=encuesta).filter(usuario__usuario__credenciales=request.user.id).count()
                        # print (Respuesta.objects.filter(pregunta__pregunta=question).filter(encuesta=encuesta).filter(usuario__usuario__credenciales=request.user.id))
                        if (count>0):
                            print ("ya se respondió la pregunta" , question)
                        else:
                            questions_filter.append(question)
                            # print (questions_filter)
        else:
            checkasign = Encuesta.objects.filter(realizar_enc__empleado__usuario__credenciales=request.user.id)
            fechahoy=str(datetime.today().strftime('%Y-%m-%d'))
            # fechahoy="2020-01-22"
            # fechahoy="2020-02-01"
            # fechahoy="2020-02-20"
            count = 0
            for check in checkasign:
                # print (check.fecha_inicio,check.fecha_termino)
                # print (fechahoy)
                if (fechahoy>=str(check.fecha_inicio) and fechahoy<=str(check.fecha_termino)):
                    count+=1
                    encuesta=check.id
                    fechafin=str(datetime.strptime(str(check.fecha_termino), '%Y-%m-%d').strftime('%d/%m/%Y'))
                    fechainicio=str(datetime.strptime(str(check.fecha_inicio), '%Y-%m-%d').strftime('%d/%m/%Y'))
                    for x in range(1,12):
                        count1 = 0
                        questions = Pregunta.objects.filter(seccion__dominio__id=x)
                        dominio_quantity.append(len(Pregunta.objects.filter(seccion__dominio__id=x)))
                        dom_description.append({'id' : x, 'name' : Dominio.objects.get(pk=x), 'description' : dominios[x-1].description})
                        for question in questions:
                            count1 += Respuesta.objects.filter(pregunta__pregunta=question).filter(encuesta=encuesta).filter(usuario__usuario__credenciales=request.user.id).count()
                            # print (" en dominio hay ppreguntas",count)
                        # print (count1)
                        # if (count < dominio_quantity[x-1]):
                        dominio_filter.append(dominio_quantity[x-1]-count1)
                        # else:
                        #     dominio_filter.append(0)
                        # print(dominio_filter)
                    for x in range (11):
                        if (dominio_filter[x]>0):
                            new_doms.append(dom_description[x])
                        else:
                            dominio_answered.append(dom_description[x])
                elif (fechahoy<str(check.fecha_inicio)):
                    encuesta_sinabrir = 1
                    fechasinabrir = check.fecha_inicio
            # si tiene alguna encuesta asociada mostrar la encuesta        
        if (count>=1):
            checksurvey=1
        else:
            checksurvey=0


        aux = Trabajador.objects.get(usuario__credenciales=request.user.id)
        if (aux.cargo.tipo_usuario.pk != 1):
            admin=1
        else:
            admin=0
        # print (admin)

        # print ("count",count)
        # print ("CHECK survey", checksurvey)

        return render(
            request,
            "users/welcome.html",
            {
                'dominio_num': dominio_num,
                'dominio_filter': dominio_filter,
                'questions': questions_filter,
                'dom_description': new_doms,
                'admin':admin,
                'checkasign':checksurvey,
                'fechainicio':fechainicio,
                'fechafin':fechafin,
                'encuesta_sinabrir':encuesta_sinabrir,
                'fechasinabrir':fechasinabrir,
                'dominio_answered':dominio_answered,
                # 'queries': queries,
                # 'dominio_num': dominio_num,
                # 'questions_quantity':questions_quantity_response,
                # 'value_answers_quantity':value_answers_quantity_response,
            })
    # En otro caso redireccionamos al login
    return redirect('/login')

def register(request):
    # Creamos el formulario de autenticación vacío
    form = UserCreationForm()
    form.fields['username'].help_text = None
    form.fields['password1'].help_text = None
    form.fields['password2'].help_text = None
    if request.method == "POST":
        # Añadimos los datos recibidos al formulario
        form = UserCreationForm(data=request.POST)
        # Si el formulario es válido...
        if form.is_valid():

            # Creamos la nueva cuenta de usuario
            user = form.save()

            # Si el usuario se crea correctamente 
            if user is not None:
                # Hacemos el login manualmente
                do_login(request, user)
                # Y le redireccionamos a la portada
                return redirect('/')

    # Si llegamos al final renderizamos el formulario
    return render(request, "users/register.html", {'form': form})
    

def login(request):
    # Creamos el formulario de autenticación vacío
    form = AuthenticationForm()
    if request.method == "POST":
        # Añadimos los datos recibidos al formulario
        form = AuthenticationForm(data=request.POST)
        # Si el formulario es válido...
        if form.is_valid():
            # Recuperamos las credenciales validadas
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']

            # Verificamos las credenciales del usuario
            user = authenticate(username=username, password=password)

            # Si existe un usuario con ese nombre y contraseña
            if user is not None:
                # Hacemos el login manualmente
                do_login(request, user)
                # Y le redireccionamos a la portada
                return redirect('/')

    # Si llegamos al final renderizamos el formulario
    return render(request, "users/login.html", {'form': form})

def logout(request):
    # Finalizamos la sesión
    do_logout(request)
    # Redireccionamos a la portada
    return redirect('/')

def myajaxtestview(request):
    getanswers = request.POST.getlist('text[]')
    # print (getanswers)
    answers=[]
    for x in getanswers:
        answers.append(x.split("-"))

    checkasign = Encuesta.objects.filter(realizar_enc__empleado__usuario__credenciales=request.user.id)
    fechahoy=str(datetime.today().strftime('%Y-%m-%d'))
    # fechahoy="2020-01-22"
    # fechahoy="2020-02-01"
    # fechahoy="2020-02-20"
    count = 0
    for check in checkasign:
        if (fechahoy>=str(check.fecha_inicio) and fechahoy<=str(check.fecha_termino)):
            encuesta=check.id
    for x in range(len(getanswers)):
        # print ("usuario ",answers[x][0])
        # print ("pregunta ",answers[x][1])
        # print ("respuesta",answers[x][2])
        respuesta = Respuesta.objects.create(pregunta=Pregunta.objects.get(pk=answers[x][1]), usuario=Trabajador.objects.get(usuario__credenciales=request.user.id), encuesta=Encuesta.objects.get(pk=encuesta), valor=int(answers[x][2]))

    return HttpResponse("Se respondieron %s preguntas" % (len(getanswers)))

def graphs (request):

    percentage_by_dom = []
    percentage_by_section = []
    percentage_unknown_dom = []
    answers_corrects = []
    answers_incorrects = []
    answers_unknown = []
    dict_sect= []
    array_encuesta = []
    respondidas_array = []
    respuestas = []
    respuestas_array = []
    usuarios_respondidos = []
    preguntas_array=[]
    encuesta_num= 0
    fechafin=0
    fechainicio=0
    restantes_por_dom=[0,0,0,0,0,0,0,0,0,0,0]
    trabajador = Trabajador.objects.get(usuario__credenciales=request.user.id)
    encuestas = Encuesta.objects.filter(supervisor=request.user.id)
    answers_per_dom = []
    answers_per_sect = []
    preguntas=[]
    fechas = []
    fechahoy = (datetime.today().strftime('%Y-%m-%d'))
    fechahoy= dateutil.parser.parse(fechahoy).date()
    respuesta_valor = []
    respuestasxdia_array=[]
    action_plan=[]
    print (len(respuesta_valor))
    aux = Trabajador.objects.get(usuario__credenciales=request.user.id)
    if (aux.cargo.tipo_usuario.pk != 1):
        admin=1
    else:
        admin=0
    for encuesta in encuestas:
        array_encuesta.append({'id': encuesta.id, 'fecha_inicio':str(datetime.strptime(str(encuesta.fecha_inicio), '%Y-%m-%d').strftime('%d/%m/%Y')), 'fecha_termino':str(datetime.strptime(str(encuesta.fecha_termino), '%Y-%m-%d').strftime('%d/%m/%Y'))})

    if request.GET.get('survey'):
        encuesta_num = request.GET.get('survey')

        if (encuesta_num != '0'):
            encuesta_escogida = Encuesta.objects.get(pk=encuesta_num)
            fechafin = encuesta_escogida.fecha_termino
            fechainicio= encuesta_escogida.fecha_inicio
            for x in Seccion.objects.all():
                aux=x.get_answers(trabajador.empresa,encuesta_num)
                dict_sect.append(aux)

            # for x in Seccion.objects.all():
            #     aux=x.get_answers()
            #     if (aux['positivo'] != 0 or aux['negativo'] != 0):
            #         dict_sect.append(aux)
            encuestas_users = Realizar_enc.objects.filter(encuesta=encuesta_num)
            for encuesta in encuestas_users:
                if (encuesta.empleado in usuarios_respondidos):
                    print ("user ya encontrado ")
                else:
                    usuarios_respondidos.append(encuesta.empleado)

            for x in range(1,12):
                answers_corrects.append(Respuesta.objects.filter(valor=1).filter(usuario__empresa__nombre_empresa=trabajador.empresa).filter(encuesta=encuesta_num).filter(pregunta__seccion__dominio_id=x).count())
                answers_incorrects.append(Respuesta.objects.filter(valor=0).filter(usuario__empresa__nombre_empresa=trabajador.empresa).filter(encuesta=encuesta_num).filter(pregunta__seccion__dominio_id=x).count())
                answers_unknown.append(Respuesta.objects.filter(valor=2).filter(usuario__empresa__nombre_empresa=trabajador.empresa).filter(encuesta=encuesta_num).filter(pregunta__seccion__dominio_id=x).count())
                answers_per_dom.append(Pregunta.objects.filter(seccion__dominio__id=x).count())

            # print (answers_corrects)
            x=0
            print (answers_per_dom)
            for dominio in Dominio.objects.all().values("nombre_dominio", "pk"):
                recomendaciones = Recomendacion.objects.filter(dominio=dominio["pk"])
                recomendaciones_array = []
                recomendaciones_name = []
                guide_array=[]
                for recomendacion in recomendaciones:
                    recomendaciones_array.append(recomendacion.descripcion)
                    recomendaciones_name.append(recomendacion.nombre)
                    guide_array.append(recomendacion.referencia)
                if (answers_unknown[x]!=0 and len(usuarios_respondidos)!=0):
                    desc=round((answers_unknown[x]/(answers_per_dom[x]*len(usuarios_respondidos)))*100)
                else:
                    desc=0
                if (answers_corrects[x]!=0 and len(usuarios_respondidos)!=0):
                    percentage_by_dom.append({'id': dominio["pk"], 'value':round((answers_corrects[x]/(answers_per_dom[x]*len(usuarios_respondidos)))*100), 'name_recomendations': recomendaciones_name, 'unknown':desc, 'recomendations': recomendaciones_array, 'guide':guide_array})
                else:
                    percentage_by_dom.append({'id': dominio["pk"], 'value':0, 'unknown':desc, 'name_recomendations': recomendaciones_name, 'recomendations': recomendaciones_array, 'guide':guide_array})
                if (percentage_by_dom[x]['value'] <= 40):
                    prioridad = 'Alta'
                    sla = '14 días'
                    slafin = (fechafin + timedelta(days=14))
                elif (percentage_by_dom[x]['value'] > 40 and percentage_by_dom[x]['value'] < 80):
                    prioridad = 'Media'
                    sla = '30 días'
                    slafin = (fechafin + timedelta(days=30))
                elif (percentage_by_dom[x]['value'] >= 80 and percentage_by_dom[x]['value'] < 100):
                    prioridad = 'Baja'
                    sla = '60 días'
                    slafin = (fechafin + timedelta(days=60))
                elif (percentage_by_dom[x]['value'] == 100):
                    prioridad = 'Nula'
                    sla = 'No existe'
                responsable = Encargado.objects.filter(dominio=Dominio.objects.get(pk=dominio["pk"]), empresa=Empresa.objects.get(nombre_empresa=trabajador.empresa))
                if not responsable:
                    responsable = "Sin asignar"
                else:
                    for user in responsable:
                        responsable = user.encargado.usuario.primer_nombre + " " + user.encargado.usuario.apellido_paterno + " " + user.encargado.usuario.apellido_materno
                acciones = Recomendacion.objects.filter(dominio = Dominio.objects.get(pk=dominio["pk"]))
                list_acciones = ""
                j=1
                for accion in acciones:
                    list_acciones += " " + str(j) + ".- " +(accion.nombre)
                    j+=1
                action_plan.append({'id': dominio["pk"], 'prioridad':prioridad, 'SLA': sla, 'responsable':responsable, 'action':list_acciones, 'initial_date': (fechafin + timedelta(days=1)), 'final_date':slafin})

                x+=1

            x=0

            for x in range(133):
                answers_per_sect.append(Pregunta.objects.filter(seccion=x+1).count())
            x=0
            # print ("dict",len(dict_sect))
            for y in dict_sect:
                # print (dict_sect[x]['id'])
                if (dict_sect[x]['positivo']!=0):
                    percentage_by_section.append({'id' : dict_sect[x]['id'], 'name' : dict_sect[x]['name'], 'dom':dict_sect[x]['dom'], 'percentage' : round(dict_sect[x]['positivo']/(answers_per_sect[x-1]*len(usuarios_respondidos))*100)})
                else:
                    percentage_by_section.append({'id' : dict_sect[x]['id'], 'name' : dict_sect[x]['name'], 'dom':dict_sect[x]['dom'], 'percentage' : 0})
                x+=1

            for x in range (len(usuarios_respondidos)):
                for y in range (1,12):
                    restantes_por_dom[y-1]=Respuesta.objects.filter(encuesta=encuesta_num).filter(usuario=usuarios_respondidos[x]).filter(pregunta__seccion__dominio_id=y).count()
                respondidas_array.append({'id':usuarios_respondidos[x].usuario, 'apellido_paterno':usuarios_respondidos[x].usuario.apellido_paterno, 'apellido_materno':usuarios_respondidos[x].usuario.apellido_materno, 'nombre': usuarios_respondidos[x].usuario.primer_nombre, 'cargo':usuarios_respondidos[x].cargo, 'Dom1': restantes_por_dom[0], 'Dom2': restantes_por_dom[1], 'Dom3': restantes_por_dom[2], 'Dom4': restantes_por_dom[3], 'Dom5': restantes_por_dom[4], 'Dom6': restantes_por_dom[5], 'Dom7': restantes_por_dom[6], 'Dom8': restantes_por_dom[7], 'Dom9': restantes_por_dom[8], 'Dom10': restantes_por_dom[9], 'Dom11': restantes_por_dom[10]})
            
            
            preguntas=Pregunta.objects.all()
            
            for pregunta in preguntas:
                preguntas_array.append({'id':pregunta.id, 'pregunta':pregunta.pregunta})

            for x in range (len(usuarios_respondidos)):
                respuesta_valor = ["","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","",""]
                respuestas=Respuesta.objects.filter(encuesta=encuesta_num).filter(usuario=usuarios_respondidos[x])
                for respuesta in respuestas:
                    if (respuesta.valor==0):
                        respuesta_valor[respuesta.pregunta.id-1]="No"
                    elif (respuesta.valor==1):
                        respuesta_valor[respuesta.pregunta.id-1]="Sí"
                    elif (respuesta.valor==2):
                        respuesta_valor[respuesta.pregunta.id-1]="Desconoce"
                respuestas_array.append({'id':x, 'nombre':usuarios_respondidos[x].usuario.primer_nombre, 'apellido_paterno':usuarios_respondidos[x].usuario.apellido_paterno, 'apellido_materno': usuarios_respondidos[x].usuario.apellido_materno, 'respuestas':respuesta_valor})

            if (str(fechahoy) < str(fechafin)):
                fechaaux2 = dateutil.parser.parse(str(fechahoy + timedelta(days=1))).date()
            else:
                fechaaux2 = (fechafin + timedelta(days=1))
            count=0
            auxiliar = abs((fechaaux2 - fechainicio).days)
            days = []
            # for x in range (auxiliar):
            #     days.append(0)
            # print (days)
            for x in range(len(usuarios_respondidos)):
                days = [0] * auxiliar
                # respuestasxdia_array.append(str(usuarios_respondidos[x].usuario.apellido_paterno) + " " + str(usuarios_respondidos[x].usuario.apellido_materno) + " " + str(usuarios_respondidos[x].usuario.primer_nombre))
                fechaaux=fechainicio
                for y in range(auxiliar):
                    days[y] = Respuesta.objects.filter(encuesta=encuesta_num).filter(usuario=usuarios_respondidos[x]).filter(fecha_respuesta=fechaaux).count()
                    if (count==0):
                        fechas.append(fechaaux)
                    fechaaux = (fechaaux + timedelta(days=1))
                contador=Respuesta.objects.filter(encuesta=encuesta_num).filter(usuario=usuarios_respondidos[x]).count()
                if ((contador/211)*100==0):
                    percentage = 0
                else:
                    percentage = format((contador/211)*100,'.2f')
                respuestasxdia_array.append({'id': x, 'nombre':usuarios_respondidos[x].usuario.primer_nombre, 'apellido_paterno':usuarios_respondidos[x].usuario.apellido_paterno, 'apellido_materno': usuarios_respondidos[x].usuario.apellido_materno, 'respuestas':days, 'percentage':percentage})
                count+=1

            
        else:
            return render(
                request,
                "users/graphs.html",
                {
                    'percentage_by_dom':percentage_by_dom,
                    'percentage_by_section':percentage_by_section,
                    'percentage_unknown':percentage_unknown_dom,
                    'empresa':trabajador.empresa,
                    'admin':admin,
                    'encuesta_selector':(array_encuesta),
                    'encuesta_number':len(array_encuesta),
                    'encuesta_picker':0,
                    'fechafin':fechafin,
                    'fechainicio':fechainicio,
                    'respondidas_array':respondidas_array,
                    'preguntas_id':preguntas_array,
                    'respuestas_array':respuestas_array,
                })

    # print (percentage_by_section)
    # print ("encuesta numm", encuesta_num)
    
    return render(
        request,
        "users/graphs.html",
        {
            'percentage_by_dom':percentage_by_dom,
            'percentage_by_section':percentage_by_section,
            'percentage_unknown':percentage_unknown_dom,
            'empresa':trabajador.empresa,
            'admin':admin,
            'encuesta_selector':(array_encuesta),
            'encuesta_number':len(array_encuesta),
            'encuesta_picker':encuesta_num,
            'fechafin':fechafin,
            'fechainicio':fechainicio,
            'respondidas_array':respondidas_array,
            'preguntas_id':preguntas_array,
            'respuestas_array':respuestas_array,
            'encuestados':len(usuarios_respondidos),
            'fechas':fechas,
            'respuestasxdia':respuestasxdia_array,
            'action_plan':action_plan,

        })
        
    # for x in range (11):
    #     sections_by_dom.append({ 'id' : x+1 , 'sections' : sections[x].nombre_seccion})
    # print ("secciones" , sections[0]['sections'])
    # aux = Respuesta.objects.all().values("pregunta__seccion__dominio__id")

    # aux = Dominio.objects.all().values("seccion__pregunta__id","pk","seccion__pregunta__respuesta").aggregate(si=Sum(Coalesce(Case(When(seccion__pregunta__respuesta__valor=1, then=1), default=0, output_field=IntegerField()), 0)))
        # dict[str(dominio["pk"])]= Sum(Coalesce(Case(When(seccion__pregunta__respuesta__valor=1, pk=dominio["pk"], then=1), default=0, output_field=IntegerField()), 0))
    # aux = Dominio.objects.all().values().annotate(nombre_dominio=F("nombre_dominio"),**dict)
    # print (aux)

    # dict = {}
    # for dominio in Dominio.objects.all().values("nombre_dominio", "pk"):
    #     answers_corrects.append(Respuesta.objects.filter(valor=1).filter(pregunta__seccion__dominio_id=dominio["pk"]).count())
    #     answers_incorrects.append(Respuesta.objects.filter(valor=0).filter(pregunta__seccion__dominio_id=dominio["pk"]).count())

def asign(request):

    supervisor = Trabajador.objects.get(usuario__credenciales=request.user.id)
    trabajadores = Trabajador.objects.filter(empresa=supervisor.empresa)

    aux = Trabajador.objects.get(usuario__credenciales=request.user.id)
    if (aux.cargo.tipo_usuario.pk != 1):
        admin=1
    else:
        admin=0

    
    return render(
        request,
        "users/registrar_enc.html",
        {
            'trabajadores':trabajadores,
            'admin':admin,
        })

def ajaxasign(request):
    datepickers = request.POST.getlist('dates[]')
    trabajadores = request.POST.getlist('users[]')
    dateinicio=datepickers[0]
    datefin=datepickers[1]
    
    msg = MIMEMultipart()
    message = "Tienes una encuesta asiganada desde el " + dateinicio + " hasta el " + datefin + ". Para contestarla dirígase a gapnuevo.herokuapp.com"

    # password ="Gap.project.2020"
    password="xjqzrodaljmqurau"
    msg['From'] = "gapproject2020@gmail.com"
    msg['Subject'] = "Nueva Encuesta Asignada"
    msg.attach(MIMEText(message, 'plain'))

    server = smtplib.SMTP('smtp.gmail.com: 587')
    server.starttls()
    server.login(msg['From'], password)

    if (dateinicio=="" or datefin==""):
        return HttpResponse("Ingrese las fechas faltantes")
    else:
        dateinicio1=time.strptime(datepickers[0],"%d/%m/%Y")
        datefin1=time.strptime(datepickers[1],"%d/%m/%Y")
    if (datefin1<dateinicio1):
        return HttpResponse("Error en las fechas ingresadas")
    else:
        dateinicio2= dateinicio[6:]+ "-" +dateinicio[3:5] + "-" + dateinicio[:2]
        datefin2= datefin[6:]+ "-" +datefin[3:5] + "-" + datefin[:2]
        encuesta = Encuesta.objects.create(supervisor=Trabajador.objects.get(usuario__credenciales=request.user.id), fecha_inicio=dateinicio2, fecha_termino=datefin2)
        print (len(trabajadores))
        supervisor = Trabajador.objects.get(usuario__credenciales=request.user.id)
        usuarios = Trabajador.objects.filter(empresa=supervisor.empresa)
        print ("USUARIOS", usuarios)
        for x in range (len(trabajadores)):
            createuser = Realizar_enc.objects.create(empleado=Trabajador.objects.get(usuario=trabajadores[x]),encuesta=Encuesta.objects.get(pk=encuesta.pk))
            if (createuser.empleado.usuario.credenciales.email != ""):
                msg['To'] = createuser.empleado.usuario.credenciales.email
                server.sendmail(msg['From'], msg['To'], msg.as_string())

        server.quit()
        return HttpResponse("Se ha creado la encuesta de manera satisfactoria")

def history(request):

    supervisor = Trabajador.objects.get(usuario__credenciales=request.user.id)
    encuestas = Encuesta.objects.filter(supervisor__empresa=supervisor.empresa)
    encuestas_array = []
    answers_corrects = [0,0,0,0,0,0,0,0,0,0,0]
    answers_incorrects = [0,0,0,0,0,0,0,0,0,0,0]
    answers_unknown = [0,0,0,0,0,0,0,0,0,0,0]
    percentage_by_dom = [0,0,0,0,0,0,0,0,0,0,0]
    x=1
    for encuesta in encuestas:
        for y in range(1,12):
            answers_corrects[y-1] = (Respuesta.objects.filter(valor=1).filter(usuario__empresa__nombre_empresa=supervisor.empresa).filter(encuesta=encuesta.pk).filter(pregunta__seccion__dominio_id=y).count())
            answers_incorrects[y-1] = (Respuesta.objects.filter(valor=0).filter(usuario__empresa__nombre_empresa=supervisor.empresa).filter(encuesta=encuesta.pk).filter(pregunta__seccion__dominio_id=y).count())
            answers_unknown[y-1] = (Respuesta.objects.filter(valor=2).filter(usuario__empresa__nombre_empresa=supervisor.empresa).filter(encuesta=encuesta.pk).filter(pregunta__seccion__dominio_id=y).count())
            print (answers_corrects[y-1], answers_incorrects[y-1], answers_unknown[y-1])
            # for dominio in Dominio.objects.all().values("nombre_dominio", "pk"):
            if (answers_corrects[y-1]!=0):
                percentage_by_dom[y-1] =(round((answers_corrects[y-1]/(answers_corrects[y-1]+answers_incorrects[y-1]+answers_unknown[y-1]))*100))
            else:
                percentage_by_dom[y-1] = (0)
            print(percentage_by_dom)
        encuestas_array.append({'id':x, 'fecha_inicio':encuesta.fecha_inicio, 'fecha_termino': encuesta.fecha_termino, 'dom1': percentage_by_dom[0], 'dom2': percentage_by_dom[1], 'dom3': percentage_by_dom[2], 'dom4': percentage_by_dom[3], 'dom5': percentage_by_dom[4], 'dom6': percentage_by_dom[5], 'dom7': percentage_by_dom[6], 'dom8': percentage_by_dom[7], 'dom9': percentage_by_dom[8], 'dom10': percentage_by_dom[9], 'dom11': percentage_by_dom[10]})
        x+=1
    print ("encuestas", encuestas_array)

    aux = Trabajador.objects.get(usuario__credenciales=request.user.id)
    if (aux.cargo.tipo_usuario.pk != 1):
        admin=1
    else:
        admin=0

    return render(
        request,
        "users/history.html",
        {
            'encuestas_array':encuestas_array,
            'admin':admin,
        })
def cumplimiento(request):

    aux = Trabajador.objects.get(usuario__credenciales=request.user.id)
    if (aux.cargo.tipo_usuario.pk != 1):
        admin=1
    else:
        admin=0

    return render(
        request,
        "users/cumplimiento.html",
        {
            'admin':admin,
        })

def ajaxxls(request):
    getsurvey= request.POST.get('text')
    encuesta = Encuesta.objects.get(pk=getsurvey)
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="users.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Users Data') # this will make a sheet named Users Data

    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['Username', 'First Name', 'Last Name', 'Email Address', ]

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style) # at 0 row 0 column 

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    rows = User.objects.all().values_list('username', 'first_name', 'last_name', 'email')
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)

    return response

def carga(request):
    aux = Trabajador.objects.get(usuario__credenciales=request.user.id)
    if (aux.cargo.tipo_usuario.pk != 1):
        admin=1
    else:
        admin=0
    return render(
    request,
    "users/carga_usuarios.html",
    {
        'admin':admin,
    })

@csrf_exempt
def ajaxload(request):
    try:
        excelfile = request.FILES['file']
        print (excelfile)
        wb_obj = load_workbook(excelfile)
        sheet_obj = wb_obj.active
        print(sheet_obj.max_row)
        trabajador = Trabajador.objects.get(usuario__credenciales=request.user.id)
        print (trabajador)
        x=0
        regex = '^\w+([\.-]?\w+)*@\w+([\.-]?\w+)*(\.\w{2,3})+$'
        msg = MIMEMultipart()
        # password ="Gap.project.2020"
        password="xjqzrodaljmqurau"
        msg['From'] = "gapproject2020@gmail.com"
        msg['Subject'] = "Usuario GAP Project"

        server = smtplib.SMTP('smtp.gmail.com: 587')
        server.starttls()
        server.login(msg['From'], password)
        for row in sheet_obj.iter_rows():
            if (x != 0):
                username = str(row[0].value[0:3]) + "." + str(row[2].value)
                message = "Se ha creado su usuario de la plataforma GAP Project con las siguientes credenciales \n usuario: " + username + " \n contraseña: Gap.project2020. \n Ingrese a gapnuevo.herokuapp.com para realizar su cambio de contraseña"
                msg.attach(MIMEText(message, 'plain'))
                usuario = User.objects.filter(username=username)
                if not usuario:
                    user = User.objects.create_user(username = username, password = 'Gap.project2020', email = row[4].value)
                    user.save()
                    #CREACION DE USUARIO
                    usuario_model = Usuario.objects.create(credenciales=User.objects.get(pk=user.pk), primer_nombre=row[0].value, segundo_nombre=row[1].value, apellido_paterno=row[2].value, apellido_materno=row[3].value, rut=row[5].value) 
                    #CREACION DE CARGO
                    cargo = Cargo.objects.filter(nombre_cargo=row[6].value)
                    if not cargo:
                        create_cargo = Cargo.objects.create(nombre_cargo=row[6].value, tipo_usuario=TipoUsuario.objects.get(nombre_tipo=row[9].value))
                    cargo = Cargo.objects.get(nombre_cargo=row[6].value)
                    #CREACION DE DPTO
                    departamento = Departamento.objects.filter(nombre_departamento=row[7].value)
                    if not departamento:
                        create_dpto = Departamento.objects.create(nombre_departamento=row[7].value)
                    departamento = Departamento.objects.get(nombre_departamento=row[7].value)
                    #CREACION DE SUCURSAL
                    sucursal = Sucursal.objects.filter(nombre_sucursal= row[8].value)
                    if not sucursal:
                        create_sucursal = Sucursal.objects.create(nombre_sucursal= row[8].value, empresa_sucursal=Empresa.objects.get(nombre_empresa= trabajador.empresa.nombre_empresa))
                    sucursal = Sucursal.objects.get(nombre_sucursal=row[8].value)
                    #CREACION DE TRABAJADOR
                    create_trabajador = Trabajador.objects.create(usuario=usuario_model, empresa=Empresa.objects.get(nombre_empresa= trabajador.empresa.nombre_empresa), cargo=cargo, dpto=departamento, sucursal=sucursal)
                    print ("trabajador", create_trabajador)
                    if (row[4].value != ""):
                        if(re.search(regex,row[4].value)):  
                            msg['To'] = row[4].value
                            server.sendmail(msg['From'], msg['To'], msg.as_string())
                        else:
                            print ("email invalido")
                    else:
                        print ("sin email")
                else:
                    print ("USUARIO "+ str(username) + " YA EXISTE")
            x+=1
        server.quit()


        return HttpResponse("Usuarios cargados exitosamente")
    except:
        return HttpResponse("Error en el archivo subido")

def responsable(request):
    aux = Trabajador.objects.get(usuario__credenciales=request.user.id)
    if (aux.cargo.tipo_usuario.pk != 1):
        admin=1
    else:
        admin=0
    trabajador = Trabajador.objects.get(usuario__credenciales=request.user.id)
    workers = Trabajador.objects.filter(empresa=Empresa.objects.get(nombre_empresa= trabajador.empresa.nombre_empresa))
    dom1 = Encargado.objects.filter(dominio=Dominio.objects.get(pk=1), empresa=Empresa.objects.get(nombre_empresa= trabajador.empresa.nombre_empresa))
    dom2 = Encargado.objects.filter(dominio=Dominio.objects.get(pk=2), empresa=Empresa.objects.get(nombre_empresa= trabajador.empresa.nombre_empresa))
    dom3 = Encargado.objects.filter(dominio=Dominio.objects.get(pk=3), empresa=Empresa.objects.get(nombre_empresa= trabajador.empresa.nombre_empresa))
    dom4 = Encargado.objects.filter(dominio=Dominio.objects.get(pk=4), empresa=Empresa.objects.get(nombre_empresa= trabajador.empresa.nombre_empresa))
    dom5 = Encargado.objects.filter(dominio=Dominio.objects.get(pk=5), empresa=Empresa.objects.get(nombre_empresa= trabajador.empresa.nombre_empresa))
    dom6 = Encargado.objects.filter(dominio=Dominio.objects.get(pk=6), empresa=Empresa.objects.get(nombre_empresa= trabajador.empresa.nombre_empresa))
    dom7 = Encargado.objects.filter(dominio=Dominio.objects.get(pk=7), empresa=Empresa.objects.get(nombre_empresa= trabajador.empresa.nombre_empresa))
    dom8 = Encargado.objects.filter(dominio=Dominio.objects.get(pk=8), empresa=Empresa.objects.get(nombre_empresa= trabajador.empresa.nombre_empresa))
    dom9 = Encargado.objects.filter(dominio=Dominio.objects.get(pk=9), empresa=Empresa.objects.get(nombre_empresa= trabajador.empresa.nombre_empresa))
    dom10 = Encargado.objects.filter(dominio=Dominio.objects.get(pk=10),  empresa=Empresa.objects.get(nombre_empresa= trabajador.empresa.nombre_empresa))
    dom11 = Encargado.objects.filter(dominio=Dominio.objects.get(pk=11),  empresa=Empresa.objects.get(nombre_empresa= trabajador.empresa.nombre_empresa))


    return render(
    request,
    "users/responsible.html",
    {
        'dom1':dom1,
        'dom2':dom2,
        'dom3':dom3,
        'dom4':dom4,
        'dom5':dom5,
        'dom6':dom6,
        'dom7':dom7,
        'dom8':dom8,
        'dom9':dom9,
        'dom10':dom10,
        'dom11':dom11,
        'workers':workers,
        'admin':admin,
    })

def ajaxresponsible(request):
    user = Trabajador.objects.get(usuario__credenciales=request.user.id)
    encargado= request.POST.get('text')
    trabajador, dominio = encargado.split("-")
    list_encargados = Encargado.objects.filter(dominio=Dominio.objects.get(pk=dominio),  empresa=Empresa.objects.get(nombre_empresa= user.empresa.nombre_empresa))
    if not list_encargados:
        create_encargado = Encargado.objects.create(encargado=Trabajador.objects.get(pk=trabajador), dominio=Dominio.objects.get(pk=dominio), empresa= Empresa.objects.get(nombre_empresa=user.empresa.nombre_empresa))
    else:
        for model in list_encargados:
            model.encargado = Trabajador.objects.get(pk=trabajador)
            model.save()

    return HttpResponse("Se ha cambiado el responsable del dominio")
